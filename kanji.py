"""Takes an excel spreadsheet containing kanji and converts it to a CSV file for use with Kotoba
Discord Bot (https://kotobaweb.com/bot). Intended to be used with Genki sheets

Starting on row 4, the excel sheet used has the following columns in this specific order:
    - Kanji number (漢字さくいん 通しNo.)
    - Textbook order (漢字読み （テキスト順）)
    - Hiragana reading (漢字の読み)
    - Kanji (漢字)
    - Unique kanji number (漢字番号.) (some kanjis have multiple readings)
    - Lesson number (課)

The resulting Kotoba CSV file has the following columns:
    - Question: Kanji
    - Answers: Hiragana
    - Comment: Meanings in English, as well as additional kunyomi/onyomi readings and lesson
      number
    - Instructions: Tells user how to answer
    - Render as: Tells Kotoba Bot how to render the kanji

The Kotoba CSV file is sorted by ascending textbook order, which also sorts unique kanji and lesson
number. This allows us to put more common readings first
"""
import sys
import csv
from openpyxl import load_workbook


def excel_to_dict(filename: str) -> dict[str, (list[str], list[int], int)]:
    """Reads rows from excel sheet and returns a dictionary of kanji

    excel sheet structure (from row 4):
        <kanji #> <textbook order> <hiragana> <kanji> <unique kanji #> <lesson #>
        int       int              str        str     int              str
    dict structure:
        { kanji: ( [<hiragana>], [<textbook order>], <lesson #> ) }
          str      list[str]     list[int]           int
    """
    # load sheet as read only
    workbook = load_workbook(filename, read_only=True)
    sheet = workbook.active

    # invariants:
    #   - len([hiragana]) == len([textbook order]) (1:1 mapping between specific hiragana and
    #     textbook order)
    #   - [<textbook order>] is sorted in ascending order and hiragana is inserted accordingly
    #       - sorting upon each insertion rather than at the end makes it easy to couple specific
    #         hiragana to textbook order
    kanji_dict = {}
    for row in sheet.iter_rows(min_row=4):
        _, text_order, hiragana, kanji, _, lesson = [cell.value for cell in row]
        if kanji not in kanji_dict:
            kanji_dict[kanji] = ([hiragana], [text_order], int(lesson[1:]))
        else:
            # if kanji already seen, add hiragana and textbook order to existing arrays
            # however, maintain sortedness of textbook order
            hiragana_list, order_list, _ = kanji_dict[kanji]
            i = _get_idx(text_order, order_list)
            hiragana_list.insert(i, hiragana)
            order_list.insert(i, text_order)

    # sort kanji by ascending textbook order (first in list)
    ordered = dict(sorted(kanji_dict.items(), key=lambda item: item[1][1][0]))
    return ordered


def _get_idx(order: int, orders: list[int]) -> int:
    """Gets index to insert current textbook order at, maintaining ascending order
    Uses linear search because of small length of list (max 6)
    """
    for i, num in enumerate(orders):
        # if current order < existing order, insert just before this
        if order < num:
            return i
    # else current order > all orders, add to end
    return len(orders)


def dict_to_csv(
    filename: str, kanji_dict: dict[str, (list[str], list[int], int)]
) -> None:
    """Takes in kanji dictionary and writes to CSV file formatted for Kotoba

    example csv structure:
        Question,Answers,Comment,Instructions,Render as
        明日,"あした,あす",Tomorrow,Type the reading!,Image
    """
    # get kanji meanings and readings from jisho(.csv) -- see jisho.py
    jisho = None
    with open("jisho.csv", "r", encoding="utf-8") as csv_file:
        reader = csv.reader(csv_file)
        jisho = {
            kanji: (meanings, kunyomi, onyomi)
            for kanji, meanings, kunyomi, onyomi in reader
        }

    # format dict for kotoba csv
    kotoba_list = [
        {
            "Question": kanji,
            "Answers": ",".join(hiragana_list),
            "Comment": f"""(L{lesson}) {jisho[kanji][0]}\nkunyomi: {jisho[kanji][1]}\nonyomi: {jisho[kanji][2]}""",
            "Instructions": "Type the reading!",
            "Render as": "Image",
        }
        for kanji, (hiragana_list, _, lesson) in kanji_dict.items()
    ]

    # write to csv
    with open(filename, "w", encoding="utf-8") as csv_file:
        fieldnames = ["Question", "Answers", "Comment", "Instructions", "Render as"]
        writer = csv.DictWriter(csv_file, fieldnames)
        writer.writeheader()
        writer.writerows(kotoba_list)


if __name__ == "__main__":
    sheet_name = "kanji.xlsx" if len(sys.argv) < 2 else sys.argv[1]
    outfile_name = "kotoba_kanji.csv" if len(sys.argv) < 3 else sys.argv[2]

    kanji_dict = excel_to_dict(sheet_name)
    dict_to_csv(outfile_name, kanji_dict)
