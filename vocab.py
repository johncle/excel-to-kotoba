"""Takes an excel spreadsheet containing general vocab and converts it to a CSV file for use with
Kotoba Discord Bot (https://kotobaweb.com/bot). Intended to be used with Genki sheets

Script takes in 3 optional arguments:
    1. Excel sheet file name (str, default 'vocab.xlsx')
    2. Output csv file name (str, default 'kotoba_vocab.csv')
    3. Duplicate (bool, default 'False')
        - True: duplicates word in each associated lesson for more accurate ranges
        - False: word appears in first associated lesson only
    4. Reverse (bool, default 'False')
        - True: Japanese to kana
        - False: English meaning to kanji or kana

Starting on row 11, the excel sheet used has the following columns in this specific order:
    - Word number (No.)
    - Word (単語) (in kana)
    - Kanji notation (漢字表記)
    - Part of speech (品詞)
        - [n.] noun [い-adj.] い-adjective [な-adj.] な-adjective
        - [u-v.] u-verb [ru-v.] ru-verb [irr-v.] irregular verb
    - English translation (英訳)
    - Lesson number (課数)
        - [会] 会話･文法編 [読] 読み書き編 [G] あいさつ [(e)] 課末コラム [I, II, III] 読み書き編の問題番号

The resulting Kotoba CSV file has the following columns:
    - Question: Kanji
    - Answers: Kana
    - Comment: Meaning(s) in English, as well as part(s) of speech and lesson number(s)
    - Instructions: Tells user how to answer
    - Render as: Tells Kotoba Bot how to render the kanji

The Kotoba CSV file is sorted by ascending lesson number
"""
import sys
import csv
import re
from openpyxl import load_workbook


def excel_to_dict(
    filename: str,
    duplicate: bool = False,
    adjustments: bool = False,
) -> dict[str, (list[str], list[str], list[str], list[str])]:
    """Reads rows from excel sheet and returns a dictionary of kanji

    excel sheet structure (from row 11):
        <word #> <kana>   <kanji> <part of speech> <english meaning> <lesson #>
        str(int) str(int) str     str              str(int)          str
        - Kanji notation may not exist for some words (purely kana)
        - Some words have multiple lesson numbers deliminated by a comma (,)
            - In this case, sort dict using the first number seen
    dict structure:
        { kanji: ( [<kana>], [<part of speech>], [<meaning>], [<lesson #>] ) }
          str      list[str] list[str]           list[str]    list[str]
        - If no kanji exists, use kana for key instead, and kana in value stays the same
    """
    # load sheet as read only
    workbook = load_workbook(filename, read_only=True)
    sheet = workbook.active

    vocab_dict = {}
    for row in sheet.iter_rows(min_row=11):
        _, kana, kanji, part, meaning, lesson = [cell.value for cell in row]
        sanitized_kana = _sanitize_kana(kana)  # list
        # if word is purely kana, use that (original kana) instead of kanji for key
        if not kanji:
            kanji = kana
        if kanji not in vocab_dict:
            vocab_dict[kanji] = (
                sanitized_kana,
                [part],
                [meaning.strip()],
                [num.strip() for num in lesson.split(",")],
            )
        else:
            # if kanji already seen, append to existing arrays
            kana_list, part_list, meaning_list, lesson_list = vocab_dict[kanji]
            # ignore duplicates
            for sanitized in sanitized_kana:
                if sanitized not in kana_list:
                    kana_list.append(sanitized.strip())
            if part not in part_list:
                part_list.append(part)
            if meaning not in meaning_list:
                meaning_list.append(meaning.strip())
            if lesson not in lesson_list:
                lesson_list.append(lesson)

    # make manual adjustments in place
    _make_adjustments(vocab_dict)
    # sort by ascending lesson number (first in lesson list)
    ordered = dict(
        sorted(vocab_dict.items(), key=lambda lesson: _lesson_sort_key(lesson[1][3][0]))
    )
    return ordered


def _sanitize_kana(kana: str) -> list[str]:
    """Some words have extra non-kana/non-katakana symbols which, at best, make it annoying to
    answer. Other times, words inside of parentheses are used for clarification which means they
    aren't even part of the reading. We remove all non-essential symbols for better user experience

    Note: '、' is interpreted by kotoba as a normal comma which separates an answer into multiple

    Examples:
        - Optional: 'おかえり（なさい）', 'ほ（う）っておく'
        - Clarification: 'いう（もんくを）', 'あんぜん（な）'
        - Tilde: 'ただの～', '～かな（あ）', 'もう～ない', '～か～'
        - Ellipse: '（～は）…といういみだ'
        - Alternatives: 'なん/なに'
        - Negative: 'あまり ＋ negative'
        - ???: '～（ん）だろう'
    """
    sanitized_kana = kana
    # remove anything within parentheses, misc characters, and ' ＋ negative'
    sanitized_kana = re.sub(r"（.*）|[～〜…！、]| ＋ negative", "", sanitized_kana)
    # split words separated by "/" (alternative readings)
    sanitized_kana = sanitized_kana.split("/")
    return sanitized_kana


def _make_adjustments(
    vocab_dict: dict[str, (list[str], list[str], list[str], list[str])],
) -> None:
    """Some entries have undesirable properties such as being in hiragana when there is a
    commonly-used kanji for it. This makes opinionated manual adjustments to those entries in place.

    - Rationale for displaying uncommonly-used kanji: It would be better to learn the uncommon kanji
      reading now than to see it and be confused later.

    Pulls from an external file (adjustments.csv):
        <kotoba line #> <original> <replacement> <answers> <comment> <split>
        str(int)        str        str           str       str       str
        - <answers>, <comment>, or <split> fields may be empty (None) to indicate no change
        - First char of comment is either 'A' (append) or 'W' (overwrite)
        - Split means to duplicate original into multiple entries for different kanjis
            - E.g. はし -> 橋 (bridge), 箸 (chopsticks)
    """
    try:
        with open("adjustments.csv", "r", encoding="utf-8") as file:
            reader = csv.reader(file)
            next(reader)  # skip header

            for entry in reader:
                _, original, replacement, answers, comment, split = entry
                # copy and remove original entry
                kana, part, meaning, lesson = vocab_dict.pop(original)
                # overwrite all answers if exist
                new_answers = answers.split(",") if answers else kana
                new_comment = meaning
                if comment:
                    # append to last comment
                    if comment[0] == "A":
                        new_comment[-1] += comment[1:]
                    # overwrite all comments
                    elif comment[0] == "W":
                        new_comment = [comment[1:]]

                # add updated entry
                if replacement in vocab_dict:
                    print(
                        "\033[93mkanji already exists:\033[0m", vocab_dict[replacement]
                    )
                vocab_dict[replacement] = (new_answers, part, new_comment, lesson)
                # bring back original if splitting
                if split:
                    vocab_dict[original] = (kana, part, meaning, lesson)
    except FileNotFoundError:
        print("\033[93madjustments.csv not found, no adjustments made\033[0m")


def _lesson_sort_key(lesson: str) -> tuple[int, int]:
    """Lesson number can have the following forms:
    - 会L## (1-2 digits)
    - 会L##(e)
    - 読L##-I, 読L##-II, 読L##-III
    - 会G
        - Greetings lesson, we interpret this as lesson 0

    We want to sort by ascending lesson number, but also need to account for 会L##(e) and 読L##-I.
    For any number x, the sort order should look as below:
        1. 会Lx
        2. 会Lx(e)
        3. 読Lx-I
        4. 読Lx-II
        5. 読Lx-III
    """
    num = "".join(c for c in lesson if c.isdigit())
    num = int(num) if num else 0  # 0 if G

    if "e" in lesson:
        rank = 1
    elif "I" in lesson:
        rank = 1 + lesson.count("I")
    else:
        rank = 0

    return num, rank


def dict_to_csv(
    filename: str, vocab_dict: dict[str, (list[str], list[str], list[str], list[str])]
) -> None:
    """Takes in vocab dictionary and writes to CSV file formatted for Kotoba

    example csv structure:
        Question,Answers,Comment,Instructions,Render as
        明日,"あした,あす",Tomorrow,Type the reading!,Image
    """
    # format dict for kotoba csv
    kotoba_list = [
        {
            "Question": kanji,
            "Answers": ",".join(kana_list),
            "Comment": f"({', '.join(lessons)}) [{', '.join(parts)}] {'; '.join(meanings)}",
            "Instructions": "Type the reading!",
            "Render as": "Image",
        }
        for kanji, (
            kana_list,
            parts,
            meanings,
            lessons,
        ) in vocab_dict.items()
    ]

    # write to csv
    with open(filename, "w", encoding="utf-8") as csv_file:
        fieldnames = ["Question", "Answers", "Comment", "Instructions", "Render as"]
        writer = csv.DictWriter(csv_file, fieldnames)
        writer.writeheader()
        writer.writerows(kotoba_list)


if __name__ == "__main__":
    sheet_name = "vocab.xlsx" if len(sys.argv) < 2 else sys.argv[1]
    outfile_name = "kotoba_vocab.csv" if len(sys.argv) < 3 else sys.argv[2]
    duplicate = len(sys.argv) > 3 and sys.argv[4].lower() == "true"
    adjustments = len(sys.argv) > 4 and sys.argv[5].lower() == "true"

    vocab = excel_to_dict(sheet_name, duplicate, adjustments)
    dict_to_csv(outfile_name, vocab)
